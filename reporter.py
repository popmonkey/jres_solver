# -*- coding: utf-8 -*-
"""
This script reads a solved race schedule from a JSON file and generates
various reports, including XLSX, CSV, and formatted text files.

:author: popmonkey and Gemini 2.5 Pro
:copyright: (c) 2025 Jules Cisek.
:license: MIT, see LICENSE for more details.
"""
import json
import datetime
import logging
import math
import argparse
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Setup Logging ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

def format_duration(duration_delta):
    """Formats a timedelta object into a human-readable string."""
    total_seconds = int(duration_delta.total_seconds())
    hours, remainder = divmod(total_seconds, 3600)
    minutes, _ = divmod(remainder, 60)
    parts = []
    if hours > 0:
        parts.append(f"{hours} hour{'s' if hours > 1 else ''}")
    if minutes > 0:
        parts.append(f"{minutes} minute{'s' if minutes > 1 else ''}")
    return "for " + " and ".join(parts) if parts else ""

def generate_member_itineraries(schedule, data, pit_time_seconds, has_spotters):
    """Generates a detailed, localized itinerary for each team member."""
    logging.info("--- Generating Per-Member Itineraries ---")
    raw_duties = {member['name']: [] for member in data['teamMembers']}
    
    for entry in schedule:
        start_time_utc = datetime.datetime.strptime(entry['startTimeUTC'], '%Y-%m-%d %H:%M:%S')
        end_time_utc = datetime.datetime.strptime(entry['endTimeUTC'], '%Y-%m-%d %H:%M:%S')
        stint_num = entry['stint']
        
        if entry['driver'] != "N/A" and entry['driver'] in raw_duties:
            raw_duties[entry['driver']].append({
                "start_utc": start_time_utc,
                "end_utc": end_time_utc,
                "activity_type": "Driving",
                "stint": stint_num
            })
        
        if has_spotters and entry['spotter'] != "N/A" and entry['spotter'] in raw_duties:
            raw_duties[entry['spotter']].append({
                "start_utc": start_time_utc,
                "end_utc": end_time_utc,
                "activity_type": "Spotting",
                "stint": stint_num
            })
            
    final_itineraries = {}
    race_start_utc = datetime.datetime.strptime(data['raceStartUTC'], "%Y-%m-%dT%H:%M:%S.%fZ")
    race_end_utc = race_start_utc + datetime.timedelta(hours=data['durationHours'])
    tz_map = {member['name']: int(member.get('timezone', 0)) for member in data['teamMembers']}

    for name, duties in raw_duties.items():
        if not duties:
            final_itineraries[name] = []
            continue
        
        duties.sort(key=lambda x: x['start_utc'])
        
        consolidated = []
        if duties:
            current_block = duties[0].copy()
            current_block['stints'] = [current_block.pop('stint')]

            for next_duty in duties[1:]:
                is_contiguous = (next_duty['activity_type'] == current_block['activity_type'] and math.isclose((next_duty['start_utc'] - current_block['end_utc']).total_seconds(), pit_time_seconds))
                if is_contiguous:
                    current_block['end_utc'] = next_duty['end_utc']
                    current_block['stints'].append(next_duty['stint'])
                else:
                    consolidated.append(current_block)
                    current_block = next_duty.copy()
                    current_block['stints'] = [current_block.pop('stint')]
            consolidated.append(current_block)

        final_itineraries[name] = []
        offset = datetime.timedelta(hours=tz_map.get(name, 0))
        last_duty_end_local = race_start_utc + offset

        for duty in consolidated:
            start_local, end_local = duty['start_utc'] + offset, duty['end_utc'] + offset
            
            gap_seconds = (start_local - last_duty_end_local).total_seconds()
            if gap_seconds > 1 and not math.isclose(gap_seconds, pit_time_seconds):
                final_itineraries[name].append({"start_local": last_duty_end_local, "end_local": start_local, "activity": "Resting"})
            
            activity_type = duty['activity_type']
            stints = duty['stints']
            if len(stints) == 1:
                activity_str = f"{activity_type} Stint #{stints[0]}"
            else:
                activity_str = f"{activity_type} Stints #{stints[0]}-{stints[-1]}"

            final_itineraries[name].append({"start_local": start_local, "end_local": end_local, "activity": activity_str})
            last_duty_end_local = end_local

        race_end_local = race_end_utc + offset
        if (race_end_local - last_duty_end_local).total_seconds() > 1:
            final_itineraries[name].append({"start_local": last_duty_end_local, "end_local": race_end_local, "activity": "Resting"})
        
    return final_itineraries

def write_report(schedule, data, filename, format):
    """Main function to generate the report in the specified format."""
    has_spotters = 'spotter' in schedule[0] if schedule else False
    
    stint_laps = int(data['fuelTankSize'] / data['fuelUsePerLap']) if data['fuelUsePerLap'] > 0 else 0
    pit_time_seconds = data['pitTimeInSeconds']
    
    final_schedule, driver_summary, spotter_summary = [], {}, {}
    driver_pool = [m for m in data['teamMembers'] if m.get('isDriver')]
    driver_summary = {p['name']: {'stints': 0, 'laps': 0} for p in driver_pool}
    
    if has_spotters:
        spotter_pool = [m for m in data['teamMembers'] if m.get('isSpotter')]
        spotter_summary = {p['name']: {'stints': 0} for p in spotter_pool}

    current_time = datetime.datetime.strptime(data['raceStartUTC'], "%Y-%m-%dT%H:%M:%S.%fZ")
    for assignment in schedule:
        start_time = current_time
        end_time = current_time + datetime.timedelta(seconds=stint_laps * data['avgLapTimeInSeconds'])
        
        entry = {
            "stint": assignment['stint'], "startTimeUTC": start_time.strftime('%Y-%m-%d %H:%M:%S'),
            "endTimeUTC": end_time.strftime('%Y-%m-%d %H:%M:%S'), "driver": assignment['driver'], "laps": stint_laps
        }
        if has_spotters:
            entry["spotter"] = assignment.get('spotter', 'N/A')
            if entry["spotter"] in spotter_summary: spotter_summary[entry["spotter"]]['stints'] += 1

        final_schedule.append(entry)
        if entry['driver'] in driver_summary:
            driver_summary[entry['driver']]['stints'] += 1
            driver_summary[entry['driver']]['laps'] += stint_laps
        
        current_time = end_time + datetime.timedelta(seconds=pit_time_seconds)
    
    member_itineraries = generate_member_itineraries(final_schedule, data, pit_time_seconds, has_spotters)

    if format == 'xlsx':
        _write_to_xlsx(final_schedule, driver_summary, spotter_summary, member_itineraries, filename)
    elif format == 'csv':
        _write_to_csv(final_schedule, filename, has_spotters)
    elif format == 'txt':
        _write_to_txt(final_schedule, driver_summary, spotter_summary, member_itineraries, filename)

def _write_to_xlsx(schedule, driver_summary, spotter_summary, member_itineraries, filename):
    """Writes all data to a multi-sheet XLSX file."""
    logging.info(f"Writing full schedule report to XLSX: {filename}")
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summaries"
    bold_font = Font(bold=True)
    
    ws_summary.cell(row=1, column=1, value="Driver Summary").font = bold_font
    ws_summary.append(["Driver", "Total Stints", "Total Laps"])
    for cell in ws_summary["2:2"]: cell.font = bold_font
    for name, stats in driver_summary.items(): ws_summary.append([name, stats['stints'], stats['laps']])
    
    if spotter_summary:
        next_row = ws_summary.max_row + 2
        ws_summary.cell(row=next_row, column=1, value="Spotter Summary").font = bold_font
        ws_summary.append(["Spotter", "Total Stints"])
        for cell in ws_summary[f"{next_row+1}:{next_row+1}"]: cell.font = bold_font
        for name, stats in spotter_summary.items():
            if stats['stints'] > 0: ws_summary.append([name, stats['stints']])

    ws_master = wb.create_sheet("Master Schedule (UTC)")
    headers = ["Stint", "Start Time (UTC)", "End Time (UTC)", "Assigned Driver"]
    if spotter_summary: headers.append("Assigned Spotter")
    headers.append("Laps")
    ws_master.append(headers)
    for cell in ws_master["1:1"]: cell.font = bold_font
    for entry in schedule:
        row = [entry['stint'], entry['startTimeUTC'], entry['endTimeUTC'], entry['driver']]
        if spotter_summary: row.append(entry['spotter'])
        row.append(entry['laps'])
        ws_master.append(row)

    # --- Per-Member Itinerary Calendar Sheets ---
    fills = {
        "Driving": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Spotting": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Resting": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    }
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for name, itinerary in member_itineraries.items():
        if not itinerary: continue
        
        ws_member = wb.create_sheet(name[:30])
        
        start_date = itinerary[0]['start_local'].date()
        end_date = itinerary[-1]['end_local'].date()
        if itinerary[-1]['end_local'].time() == datetime.time(0, 0):
            end_date -= datetime.timedelta(days=1)
        
        date_range = pd.date_range(start_date, end_date)
        
        columns = [date.strftime('%Y-%m-%d') for date in date_range]
        index = [f"{h:02d}:{m:02d}" for h in range(24) for m in range(0, 60, 15)]
        df = pd.DataFrame(index=index, columns=columns)

        for date_str in df.columns:
            for time_str in df.index:
                slot_start_time = datetime.datetime.strptime(f"{date_str} {time_str}", '%Y-%m-%d %H:%M')
                slot_end_time = slot_start_time + datetime.timedelta(minutes=15)
                
                activity_for_slot = "Resting"
                for duty in itinerary:
                    if duty['start_local'] < slot_end_time and duty['end_local'] > slot_start_time:
                        activity_for_slot = duty['activity']
                        break
                
                df.loc[time_str, date_str] = activity_for_slot
        
        ws_member.cell(row=1, column=1, value=f"Schedule for {name}").font = bold_font
        ws_member.cell(row=2, column=1, value="Time (Local)").font = bold_font
        for c_idx, col_name in enumerate(df.columns, 2):
            ws_member.cell(row=2, column=c_idx, value=col_name).font = bold_font
        
        for r_idx, row_name in enumerate(df.index, 3):
            ws_member.cell(row=r_idx, column=1, value=row_name)

        for r_idx, row in enumerate(df.values, 3):
            for c_idx, value in enumerate(row, 2):
                cell = ws_member.cell(row=r_idx, column=c_idx, value=str(value) if pd.notna(value) else "")
                activity_type = str(value).split(' ')[0]
                cell.fill = fills.get(activity_type, fills["Resting"])
                cell.border = thin_border
                cell.alignment = center_align

    for sheet in wb.worksheets:
        for col in sheet.columns:
            sheet.column_dimensions[get_column_letter(col[0].column)].width = 25
    
    wb.save(filename)

def _write_to_csv(schedule, filename, has_spotters):
    """Writes the master schedule to a CSV file."""
    logging.info(f"Writing master schedule to CSV: {filename}")
    headers = ["Stint", "Start Time (UTC)", "End Time (UTC)", "Assigned Driver"]
    if has_spotters: headers.append("Assigned Spotter")
    headers.append("Laps")
    with open(filename, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for entry in schedule:
            row = [entry['stint'], entry['startTimeUTC'], entry['endTimeUTC'], entry['driver']]
            if has_spotters: row.append(entry['spotter'])
            row.append(entry['laps'])
            writer.writerow(row)

def _write_to_txt(schedule, driver_summary, spotter_summary, member_itineraries, filename):
    """Writes all schedule data to a text file."""
    logging.info(f"Generating TXT report to file: {filename}")
    with open(filename, 'w') as f:
        f.write("--- DRIVER SUMMARY ---\n" + "="*80 + "\n")
        f.write(f"{'Driver':<20} | {'Total Stints':<15} | {'Total Laps':<15}\n" + "-"*80 + "\n")
        for name, stats in driver_summary.items(): f.write(f"{name:<20} | {stats['stints']:<15} | {stats['laps']:<15}\n")

        if spotter_summary:
            f.write("\n--- SPOTTER SUMMARY ---\n" + "="*80 + "\n")
            f.write(f"{'Spotter':<20} | {'Total Stints':<15}\n" + "-"*80 + "\n")
            for name, stats in spotter_summary.items():
                if stats['stints'] > 0: f.write(f"{name:<20} | {stats['stints']:<15}\n")
        
        f.write("\n--- MASTER SCHEDULE (UTC) ---\n" + "="*80 + "\n")
        headers = f"{'Stint':<7} | {'Start Time (UTC)':<22} | {'End Time (UTC)':<22} | {'Driver':<15}"
        if spotter_summary: headers += f" | {'Spotter':<15}"
        f.write(headers + "\n")
        f.write("-" * (len(headers) + 10) + "\n")
        for entry in schedule:
            row = f"{entry['stint']:<7} | {entry['startTimeUTC']:<22} | {entry['endTimeUTC']:<22} | {entry['driver']:<15}"
            if spotter_summary: row += f" | {entry['spotter']:<15}"
            f.write(row + "\n")

        f.write("\n--- MEMBER ITINERARIES (LOCAL TIME) ---\n" + "="*80 + "\n")
        for name, itinerary in member_itineraries.items():
            if not itinerary: continue
            f.write(f"\n--- Itinerary for {name} ---\n")
            for duty in itinerary:
                f.write(f"  {duty['start_local'].strftime('%Y-%m-%d %H:%M')} to {duty['end_local'].strftime('%H:%M')} -> {duty['activity']} {format_duration(duty['end_local'] - duty['start_local'])}\n")

def main():
    """Main function to parse arguments, read data, and generate reports."""
    parser = argparse.ArgumentParser(description="Generate reports from a solved race schedule.")
    parser.add_argument('input_file', help="Path to the solved_schedule.json file.")
    parser.add_argument('output_file', help="Path to save the report file.")
    parser.add_argument('--format', choices=['xlsx', 'csv', 'txt'], default='xlsx', help="Output format for the report.")
    args = parser.parse_args()

    try:
        with open(args.input_file, 'r') as f:
            solved_data = json.load(f)
        write_report(solved_data['schedule'], solved_data['raceData'], args.output_file, args.format)
    except Exception as e:
        logging.error(f"Failed to generate report: {e}", exc_info=True)

if __name__ == '__main__':
    main()
